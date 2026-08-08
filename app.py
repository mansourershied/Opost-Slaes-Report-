from __future__ import annotations

import json
import base64
import io
import os
import sqlite3
import sys
import threading
import time
import secrets
import shutil
from concurrent.futures import ThreadPoolExecutor
import uuid
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from report_service import generate_monthly_report
from business_service import field_value
from business_profile_service import (
    build_business_profile,
    search_businesses,
)
from analytics_service import load_snapshot, refresh_in_background, refresh_now, refresh_state
from opost_client import (
    OpostClient,
    opost_list_account_managers,
    opost_find_manager_candidates,
    opost_bulk_change_account_manager,
)

def _resolve_user_data_dir() -> Path:
    """Return a version-independent local data folder on Windows.

    This keeps created site accounts when a new project ZIP/version replaces the
    previous folder. OPOST_USER_DATA_DIR can override the location explicitly.
    """
    override = str(os.getenv("OPOST_USER_DATA_DIR") or "").strip()
    if override:
        return Path(override).expanduser().resolve()
    local_app_data = str(os.getenv("LOCALAPPDATA") or "").strip()
    if os.name == "nt" and local_app_data:
        return Path(local_app_data) / "Opost-Slaes-Report" / "data"
    return PROJECT_DIR / "data"


USER_DATA_DIR = _resolve_user_data_dir()
USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
DATABASE_PATH = USER_DATA_DIR / "employees.db"
SEED_DATABASE_PATH = PROJECT_DIR / "data" / "employees.db"
LEGACY_DATABASE_PATH = BASE_DIR / "employees.db"
USERS_BACKUP_PATH = USER_DATA_DIR / "users_backup.json"
FOLLOW_UP_DATA_PATH = BASE_DIR / "latest_follow_up.json"
DASHBOARD_DATA_PATH = BASE_DIR / "latest_dashboard.json"
JOBS_DIR = BASE_DIR / "jobs"
JOBS_DIR.mkdir(exist_ok=True)
IDENTITY_DIR = PROJECT_DIR / "secure_data" / "business_identities"
IDENTITY_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_IDENTITY_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
IDENTITY_SETTINGS_PATH = USER_DATA_DIR / "identity_settings.json"
ACCOUNT_MANAGERS_CACHE_PATH = USER_DATA_DIR / "opost_account_managers_cache.json"
DOWNLOADS_DIR = USER_DATA_DIR / "download_center"
DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
ACCOUNT_MANAGERS_CACHE_SECONDS = 30 * 60
_ACCOUNT_MANAGERS_LOCK = threading.RLock()

SOURCE_ACCOUNT_MANAGER = "osama abuqweider (sales)"
TARGET_ACCOUNT_MANAGER = "خدمة المشتركين - 0568823212"


app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    MAX_CONTENT_LENGTH=100 * 1024 * 1024,
)

DEFAULT_ADMIN_NAME = "Mansour Ershied"
DEFAULT_ADMIN_EMAIL = "Mansour_E@gmail.com"
DEFAULT_ADMIN_PASSWORD = "mansour2007"

_JOB_LOCK = threading.Lock()
_REPORT_EXECUTOR = ThreadPoolExecutor(max_workers=max(2, min(4, int(os.getenv("REPORT_WORKERS", "4")))), thread_name_prefix="opost-report")
_ACTIVE_REPORTS: dict[str, str] = {}
_ACTIVE_REPORTS_LOCK = threading.RLock()
_IDENTITY_SYNC_EXECUTOR = ThreadPoolExecutor(max_workers=1, thread_name_prefix="identity-sync")
_IDENTITY_SYNC_LOCK = threading.RLock()
_IDENTITY_SYNC_ACTIVE = False


class ReportCancelled(Exception):
    """Raised when the owner requests that a running report stops."""




def migrate_legacy_database() -> None:
    """Recover the newest existing user database into the persistent data folder.

    New ZIP versions are often extracted to a new sibling directory. On Windows,
    user accounts therefore live under LOCALAPPDATA and the first run copies the
    newest existing database it can safely find.
    """
    if DATABASE_PATH.exists():
        return

    candidates: list[Path] = []
    for candidate in (SEED_DATABASE_PATH, LEGACY_DATABASE_PATH):
        if candidate.exists() and candidate.resolve() != DATABASE_PATH.resolve():
            candidates.append(candidate)

    try:
        project_parent = PROJECT_DIR.parent
        for folder in project_parent.iterdir():
            if not folder.is_dir() or folder.resolve() == PROJECT_DIR.resolve():
                continue
            name = folder.name.lower()
            if "opost" not in name or "report" not in name:
                continue
            for candidate in (folder / "data" / "employees.db", folder / "web_app" / "employees.db"):
                if candidate.exists():
                    candidates.append(candidate)
    except OSError:
        pass

    if candidates:
        try:
            newest = max(candidates, key=lambda item: item.stat().st_mtime)
            DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(newest, DATABASE_PATH)
            print(f"Recovered site users database from: {newest}")
        except OSError as error:
            print(f"User database migration skipped: {error}")


def backup_users() -> None:
    """Write a readable backup whenever site accounts change."""
    try:
        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        rows = connection.execute(
            "SELECT id,full_name,email,role,active,can_manage_identities,can_transfer_account_managers,can_view_analytics,can_manage_users,created_at FROM employees ORDER BY id"
        ).fetchall()
        connection.close()
        write_json(USERS_BACKUP_PATH, [dict(row) for row in rows])
    except Exception:
        pass


def get_database() -> sqlite3.Connection:
    # WAL lets readers continue while another request is writing.  The busy
    # timeout prevents short concurrent writes from failing immediately.
    connection = sqlite3.connect(DATABASE_PATH, timeout=10.0)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA busy_timeout=10000")
    connection.execute("PRAGMA foreign_keys=ON")
    return connection


def initialize_database() -> None:
    connection = get_database()
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'employee',
            active INTEGER NOT NULL DEFAULT 1,
            can_manage_identities INTEGER NOT NULL DEFAULT 0,
            can_transfer_account_managers INTEGER NOT NULL DEFAULT 0,
            can_view_analytics INTEGER NOT NULL DEFAULT 0,
            can_manage_users INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    columns = {row[1] for row in connection.execute("PRAGMA table_info(employees)").fetchall()}
    if "can_manage_identities" not in columns:
        connection.execute("ALTER TABLE employees ADD COLUMN can_manage_identities INTEGER NOT NULL DEFAULT 0")
    if "can_transfer_account_managers" not in columns:
        connection.execute("ALTER TABLE employees ADD COLUMN can_transfer_account_managers INTEGER NOT NULL DEFAULT 0")
    if "can_view_analytics" not in columns:
        connection.execute("ALTER TABLE employees ADD COLUMN can_view_analytics INTEGER NOT NULL DEFAULT 0")
    if "can_manage_users" not in columns:
        connection.execute("ALTER TABLE employees ADD COLUMN can_manage_users INTEGER NOT NULL DEFAULT 0")
    connection.execute(
        """CREATE TABLE IF NOT EXISTS business_identities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            business_id TEXT NOT NULL UNIQUE,
            business_name TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            uploaded_by INTEGER NOT NULL,
            uploaded_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS identity_sync_files (
            source_path TEXT PRIMARY KEY,
            file_mtime REAL NOT NULL,
            file_size INTEGER NOT NULL,
            file_hash TEXT NOT NULL,
            business_id TEXT,
            synced_at TEXT NOT NULL
        )"""
    )
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(active)")
    connection.commit()
    connection.close()


def create_default_admin() -> None:
    """Guarantee the built-in administrator account is usable.

    The account Mansour_E@gmail.com / mansour2007 is the recovery administrator for
    this installation. Passwords are stored only as Werkzeug hashes in SQLite.
    Other users are never reset or recreated here.
    """
    connection = get_database()
    row = connection.execute(
        "SELECT id,password_hash FROM employees WHERE email = ? COLLATE BINARY",
        (DEFAULT_ADMIN_EMAIL,),
    ).fetchone()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not row:
        connection.execute(
            """INSERT INTO employees
               (full_name,email,password_hash,role,active,can_manage_identities,
                can_transfer_account_managers,can_view_analytics,can_manage_users,created_at)
               VALUES (?,?,?,'admin',1,1,1,1,1,?)""",
            (DEFAULT_ADMIN_NAME, DEFAULT_ADMIN_EMAIL, generate_password_hash(DEFAULT_ADMIN_PASSWORD), now),
        )
    else:
        # This is the one fixed recovery account requested for the project. Repair
        # its credentials/role if an older ZIP or database left it inconsistent.
        needs_password_repair = not check_password_hash(row["password_hash"], DEFAULT_ADMIN_PASSWORD)
        if needs_password_repair:
            connection.execute(
                "UPDATE employees SET password_hash=? WHERE id=?",
                (generate_password_hash(DEFAULT_ADMIN_PASSWORD), row["id"]),
            )
        connection.execute(
            """UPDATE employees SET full_name=?, role='admin', active=1,
               can_manage_identities=1, can_transfer_account_managers=1,
               can_view_analytics=1, can_manage_users=1
               WHERE id=?""",
            (DEFAULT_ADMIN_NAME, row["id"]),
        )
    connection.commit()
    connection.close()
    backup_users()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "employee_id" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        employee_id = session.get("employee_id")
        if not employee_id:
            return redirect(url_for("login"))

        # Read the role from the database on every protected request. This avoids
        # stale session roles after an administrator account is repaired or edited.
        connection = get_database()
        employee = connection.execute(
            "SELECT full_name, role, active FROM employees WHERE id=?",
            (employee_id,),
        ).fetchone()
        connection.close()

        if not employee or employee["active"] != 1:
            session.clear()
            return redirect(url_for("login"))

        role = str(employee["role"] or "").strip().lower()
        session["employee_role"] = role
        session["employee_name"] = employee["full_name"]
        if role != "admin":
            flash("هذه الصفحة متاحة للمدير فقط.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def get_current_employee():
    employee_id = session.get("employee_id")
    if not employee_id:
        return None
    connection = get_database()
    employee = connection.execute(
        "SELECT id,full_name,email,role,active,can_manage_identities,can_transfer_account_managers,can_view_analytics,can_manage_users FROM employees WHERE id=?",
        (employee_id,),
    ).fetchone()
    connection.close()
    return employee


def can_manage_identities() -> bool:
    employee = get_current_employee()
    return bool(employee and (str(employee["role"]).lower() == "admin" or int(employee["can_manage_identities"] or 0) == 1))


def can_transfer_account_managers() -> bool:
    employee = get_current_employee()
    return bool(employee and (str(employee["role"]).lower() == "admin" or int(employee["can_transfer_account_managers"] or 0) == 1))


def can_view_analytics() -> bool:
    employee = get_current_employee()
    return bool(employee and (str(employee["role"]).lower() == "admin" or int(employee["can_view_analytics"] or 0) == 1))


def can_manage_users() -> bool:
    employee = get_current_employee()
    return bool(employee and (str(employee["role"]).lower() == "admin" or int(employee["can_manage_users"] or 0) == 1))


def users_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "employee_id" not in session:
            return redirect(url_for("login"))
        if not can_manage_users():
            flash("لا تملك صلاحية إدارة مستخدمي الموقع.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def analytics_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "employee_id" not in session:
            return redirect(url_for("login"))
        if not can_view_analytics():
            flash("لا تملك صلاحية الوصول إلى الإحصائيات.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def account_manager_transfer_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "employee_id" not in session:
            return redirect(url_for("login"))
        if not can_transfer_account_managers():
            flash("لا تملك صلاحية تحويل مدير الحساب.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def identities_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if "employee_id" not in session:
            return redirect(url_for("login"))
        if not can_manage_identities():
            flash("لا تملك صلاحية إدارة هويات الحسابات.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def normalize_business_name(value: str) -> str:
    import re
    text = str(value or "").strip().casefold()
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ى", "ي").replace("ة", "ه")
    return re.sub(r"[^\w\u0600-\u06ff]+", "", text)


def get_identity_record(business_id: str):
    connection = get_database()
    row = connection.execute("SELECT * FROM business_identities WHERE business_id=?", (str(business_id),)).fetchone()
    connection.close()
    return row


@app.context_processor
def inject_template_data():
    return {
        "current_employee": get_current_employee(),
        "current_year": datetime.now().year,
        "can_manage_identities": can_manage_identities(),
        "can_transfer_account_managers": can_transfer_account_managers(),
        "can_view_analytics": can_view_analytics(),
        "can_manage_users": can_manage_users(),
    }


def read_json(path: Path, default: Any):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        pass
    return default


def write_json(path: Path, payload: Any) -> None:
    """Write job state safely even when OneDrive/antivirus briefly locks files."""
    import os
    import random

    path.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    last_error: OSError | None = None

    # Use a unique temporary name so simultaneous readers/writers never fight
    # over one shared .tmp file. Windows/OneDrive may briefly lock either file,
    # therefore replacement is retried with a short backoff.
    for attempt in range(8):
        temp = path.with_name(
            f".{path.name}.{os.getpid()}.{threading.get_ident()}.{random.randrange(1_000_000):06d}.tmp"
        )
        try:
            with temp.open("w", encoding="utf-8", newline="") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temp, path)
            return
        except (PermissionError, OSError) as exc:
            last_error = exc
            try:
                temp.unlink(missing_ok=True)
            except OSError:
                pass
            time.sleep(0.05 * (attempt + 1))

    # Last-resort in-place write. This avoids failing an otherwise completed
    # employee report solely because OneDrive held the destination momentarily.
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        return
    except OSError:
        if last_error is not None:
            raise last_error
        raise


def _unique_download_path(filename: str) -> Path:
    """Return a safe persistent path without overwriting an older file."""
    safe = secure_filename(filename) or f"download-{datetime.now():%Y%m%d-%H%M%S}"
    target = DOWNLOADS_DIR / safe
    if not target.exists():
        return target
    stem, suffix = target.stem, target.suffix
    for index in range(2, 10000):
        candidate = DOWNLOADS_DIR / f"{stem}-{index}{suffix}"
        if not candidate.exists():
            return candidate
    return DOWNLOADS_DIR / f"{stem}-{uuid.uuid4().hex[:8]}{suffix}"


def _archive_generated_file(source: Path, preferred_name: str) -> Path:
    """Copy a generated file into Download Center while preserving the original."""
    target = _unique_download_path(preferred_name)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return target


def _format_file_size(size: int) -> str:
    value = float(max(0, size))
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{size} B"


def job_path(job_id: str) -> Path:
    return JOBS_DIR / f"{job_id}.json"


def save_job(job_id: str, **changes: Any) -> dict[str, Any]:
    with _JOB_LOCK:
        data = read_json(job_path(job_id), {"id": job_id})
        data.update(changes)
        data["updated_at"] = datetime.now().isoformat(timespec="seconds")
        write_json(job_path(job_id), data)
        return data


def _report_key(form: dict[str, str], owner_id: int) -> str:
    parts = [str(owner_id), form.get("start_date", ""), form.get("end_date", ""), form.get("minimum_delivery", "30"), form.get("minimum_shipments", "30")]
    return "|".join(parts)

def _release_active_report(job_id: str) -> None:
    with _ACTIVE_REPORTS_LOCK:
        for key, value in list(_ACTIVE_REPORTS.items()):
            if value == job_id:
                _ACTIVE_REPORTS.pop(key, None)

def run_report_job(job_id: str, form: dict[str, str]) -> None:
    started_monotonic = time.monotonic()
    started_monotonic_epoch = time.time()
    maximum_seconds = None

    def update_progress(progress: int, stage: str, details: dict[str, Any]) -> None:
        current = read_json(job_path(job_id), {})
        if current.get("cancel_requested"):
            raise ReportCancelled("تم إيقاف إنشاء التقرير بواسطة المستخدم.")
        elapsed = max(0.0, time.monotonic() - started_monotonic)
        remaining = None
        if progress > 5 and progress < 100:
            remaining = max(0, int((elapsed / progress) * (100 - progress)))
        save_job(
            job_id,
            status="running",
            progress=progress,
            stage=stage,
            elapsed_seconds=int(elapsed),
            estimated_remaining_seconds=remaining,
            maximum_seconds=None,
            started_at_epoch=started_monotonic_epoch,
            completed_items=details.get("completed"),
            total_items=details.get("total") or details.get("accounts"),
        )

    try:
        update_progress(3, "بدء إنشاء التقرير", {})
        result = generate_monthly_report(
            start_date=form["start_date"],
            end_date=form["end_date"],
            minimum_delivery=form.get("minimum_delivery", "30"),
            minimum_shipments=form.get("minimum_shipments", "30"),
            return_percentage="10",
            incubation_days="40",
            progress_callback=update_progress,
        )
        current_after_generate = read_json(job_path(job_id), {})
        if current_after_generate.get("cancel_requested") or current_after_generate.get("status") == "cancelled":
            raise ReportCancelled("تم إيقاف إنشاء التقرير بواسطة المستخدم.")
        if not result.get("success"):
            raise RuntimeError(result.get("error") or "تعذر إنشاء التقرير")

        summary = result.get("summary", {})
        follow_up_accounts = result.get("follow_up_accounts", [])
        write_json(FOLLOW_UP_DATA_PATH, follow_up_accounts)
        dashboard_data = {
            "summary": summary,
            "start_date": result.get("start_date", form["start_date"]),
            "end_date": result.get("end_date", form["end_date"]),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source": "OPOST",
        }
        write_json(DASHBOARD_DATA_PATH, dashboard_data)
        source_report_path = Path(result["file_path"]).resolve()
        preferred_report_name = report_download_name(form["start_date"], form["end_date"])
        try:
            archived_report_path = _archive_generated_file(source_report_path, preferred_report_name)
        except Exception:
            archived_report_path = source_report_path
        file_path = str(archived_report_path)
        elapsed = int(time.monotonic() - started_monotonic)
        save_job(
            job_id,
            status="completed",
            progress=100,
            stage="اكتمل التقرير",
            summary=summary,
            file_path=file_path,
            elapsed_seconds=elapsed,
            estimated_remaining_seconds=0,
            completed_at=datetime.now().isoformat(timespec="seconds"),
        )
    except ReportCancelled as error:
        elapsed = int(time.monotonic() - started_monotonic)
        save_job(
            job_id,
            status="cancelled",
            progress=100,
            stage="تم إيقاف التقرير",
            elapsed_seconds=elapsed,
            estimated_remaining_seconds=0,
            error=str(error),
        )
    except Exception as error:
        elapsed = int(time.monotonic() - started_monotonic)
        save_job(
            job_id,
            status="failed",
            progress=100,
            stage="فشل إنشاء التقرير",
            elapsed_seconds=elapsed,
            estimated_remaining_seconds=0,
            error=str(error),
        )
    finally:
        _release_active_report(job_id)


def start_report_job(form: dict[str, str], owner_id: int) -> dict[str, Any]:
    key = _report_key(form, owner_id)
    with _ACTIVE_REPORTS_LOCK:
        existing_id = _ACTIVE_REPORTS.get(key)
        if existing_id:
            existing = read_json(job_path(existing_id), None)
            if existing and existing.get("status") in {"queued", "running"}:
                return existing
            _ACTIVE_REPORTS.pop(key, None)
        job_id = uuid.uuid4().hex[:12]
        _ACTIVE_REPORTS[key] = job_id
    payload = {
        "id": job_id,
        "owner_id": owner_id,
        "status": "queued",
        "progress": 2,
        "stage": "تمت إضافة التقرير إلى قائمة التنفيذ",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "start_date": form["start_date"],
        "end_date": form["end_date"],
        "minimum_delivery": form.get("minimum_delivery", "30"),
        "minimum_shipments": form.get("minimum_shipments", "30"),
        "maximum_return": "10",
        "incubation_days": "40",
        "elapsed_seconds": 0,
        "estimated_remaining_seconds": None,
        "maximum_seconds": None,
        "started_at_epoch": time.time(),
        "cancel_requested": False,
    }
    write_json(job_path(job_id), payload)
    _REPORT_EXECUTOR.submit(run_report_job, job_id, form)
    return payload


@app.route("/")
def index():
    # Always start from a clean login screen. The default administrator exists
    # in the database, but credentials are never prefilled or auto-submitted.
    session.clear()
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        # Login/logout is isolated in the current browser tab by sessionStorage.
        # Do not clear the shared cookie here, otherwise every open tab exits.
        pass
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        connection = get_database()
        employee = connection.execute("SELECT * FROM employees WHERE email = ? COLLATE BINARY", (email,)).fetchone()
        connection.close()
        if not employee:
            flash("البريد الإلكتروني غير موجود. تحقق من البريد وحاول مرة أخرى.", "error")
            return render_template("login.html", entered_email=email)
        if int(employee["active"] or 0) != 1:
            flash("هذا الحساب موقوف. راجع مدير النظام.", "error")
            return render_template("login.html", entered_email=email)
        if not check_password_hash(employee["password_hash"], password):
            flash("كلمة المرور غير صحيحة.", "error")
            return render_template("login.html", entered_email=email)

        session.clear()
        session.update(
            employee_id=employee["id"],
            employee_name=employee["full_name"],
            employee_role=str(employee["role"] or "").strip().lower(),
        )
        return redirect(url_for("dashboard"))
    return render_template("login.html", entered_email="")


@app.route("/logout")
def logout():
    # Soft per-tab logout. Other open tabs keep their current work/session.
    return redirect(url_for("login", tab_logout="1"))


@app.route("/dashboard")
@login_required
def dashboard():
    dashboard_data = read_json(DASHBOARD_DATA_PATH, {})
    summary = dashboard_data.get("summary", {})
    cards = {
        "accounts_created": summary.get("total_accounts"),
        "accounts_with_shipments": summary.get("accounts_with_shipments"),
        "accounts_without_shipments": summary.get("accounts_without_shipments"),
        "best_accounts": summary.get("best_accounts"),
        "need_follow_up": summary.get("follow_up_accounts"),
    }
    report_period = "لا يوجد تحديث ناجح حتى الآن"
    if dashboard_data.get("start_date") and dashboard_data.get("end_date"):
        report_period = f'{dashboard_data["start_date"]} ← {dashboard_data["end_date"]}'
    hour = datetime.now().hour
    greeting_type = "morning" if 5 <= hour < 12 else "evening"
    return render_template(
        "dashboard.html",
        cards=cards,
        greeting_type=greeting_type,
        report_period=report_period,
        last_updated=dashboard_data.get("updated_at", ""),
    )


@app.route("/business-profile", methods=["GET", "POST"])
@login_required
def business_profile():
    query = request.values.get("query", "").strip()
    business_id = request.values.get("business_id", "").strip()
    results = []
    profile = None
    error = ""
    try:
        if business_id:
            profile = build_business_profile(business_id)
            identity = get_identity_record(business_id) if can_manage_identities() else None
            if profile is not None:
                profile["identity"] = dict(identity) if identity else None
        elif query:
            results = search_businesses(query)
            if not results:
                error = "لم يتم العثور على حساب مطابق في OPOST."
    except Exception as exc:
        error = str(exc)
    return render_template(
        "business_profile.html",
        query=query,
        results=results,
        profile=profile,
        error=error,
    )


@app.get("/business-profile/pdf")
@login_required
def business_profile_pdf():
    business_id = request.args.get("business_id", "").strip()
    if not business_id:
        abort(400)
    theme = "site" if request.args.get("theme") == "site" else "official"
    include_identity = request.args.get("include_identity", "1") == "1"
    profile = build_business_profile(business_id)
    identity_data_uri = ""
    if include_identity and can_manage_identities():
        row = get_identity_record(business_id)
        if row:
            path = IDENTITY_DIR / str(business_id) / row["stored_filename"]
            if path.exists():
                mime = {".png":"image/png", ".webp":"image/webp"}.get(path.suffix.lower(), "image/jpeg")
                identity_data_uri = f"data:{mime};base64," + base64.b64encode(path.read_bytes()).decode("ascii")
    language = request.args.get("language", request.cookies.get("language", "ar"))
    language = "en" if language == "en" else "ar"
    html = render_template(
        "business_profile_pdf.html",
        profile=profile,
        identity_data_uri=identity_data_uri,
        theme=theme,
        language=language,
    )
    from playwright.sync_api import sync_playwright
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1240, "height": 1754})
        page.set_content(html, wait_until="load")
        page.emulate_media(media="print")
        pdf_bytes = page.pdf(
            format="A4",
            print_background=True,
            margin={"top":"0", "right":"0", "bottom":"0", "left":"0"},
            prefer_css_page_size=True,
        )
        browser.close()
    filename = f"Business Profile - {profile.get('row', {}).get('Business Name', business_id)}.pdf"
    try:
        archived_pdf = _unique_download_path(filename)
        archived_pdf.write_bytes(pdf_bytes)
    except Exception:
        archived_pdf = None
    return send_file(
        archived_pdf if archived_pdf and archived_pdf.exists() else io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )


@app.post("/api/dashboard/refresh")
@login_required
def dashboard_refresh():
    now = datetime.now()
    start_date = now.replace(day=1).strftime("%Y-%m-%d")
    end_date = now.strftime("%Y-%m-%d")
    form = {"start_date": start_date, "end_date": end_date, "minimum_delivery": "30", "minimum_shipments": "30", "maximum_return": "10", "incubation_days": "40"}
    job = start_report_job(form, int(session["employee_id"]))
    return jsonify(job), 202


@app.route("/reports/monthly", methods=["GET", "POST"])
@login_required
def monthly_report():
    if request.method == "POST":
        form = {
            "start_date": request.form.get("start_date", "").strip(),
            "end_date": request.form.get("end_date", "").strip(),
            "minimum_delivery": request.form.get("minimum_delivery", "30").strip(),
            "minimum_shipments": request.form.get("minimum_shipments", "30").strip(),
            "maximum_return": "10",
            "incubation_days": "40",
        }
        try:
            start_value = datetime.strptime(form["start_date"], "%Y-%m-%d")
            end_value = datetime.strptime(form["end_date"], "%Y-%m-%d")
            if start_value > end_value:
                raise ValueError("تاريخ البداية يجب أن يكون قبل تاريخ النهاية")
        except ValueError as error:
            flash(str(error) if form["start_date"] and form["end_date"] else "اختر فترة التقرير.", "error")
            return redirect(url_for("monthly_report"))
        job = start_report_job(form, int(session["employee_id"]))
        return redirect(url_for("monthly_report", job=job["id"]))
    return render_template("monthly_report.html", selected_job=request.args.get("job", ""))


def _job_access_allowed(data: dict[str, Any]) -> bool:
    try:
        owner_id = int(data.get("owner_id"))
        employee_id = int(session.get("employee_id"))
    except (TypeError, ValueError):
        return False
    role = str(session.get("employee_role") or "").strip().lower()
    return owner_id == employee_id or role == "admin"


@app.get("/api/jobs/<job_id>")
@login_required
def job_status(job_id: str):
    data = read_json(job_path(job_id), None)
    if not data:
        return jsonify({"error": "المهمة غير موجودة"}), 404
    if not _job_access_allowed(data):
        return jsonify({"error": "المهمة تخص جلسة دخول أخرى"}), 403
    safe = dict(data)
    safe.pop("file_path", None)
    safe["download_url"] = url_for("download_job", job_id=job_id) if data.get("status") == "completed" else None
    return jsonify(safe)




@app.post("/api/jobs/<job_id>/cancel")
@login_required
def cancel_job(job_id: str):
    data = read_json(job_path(job_id), None)
    if not data:
        return jsonify({"error": "المهمة غير موجودة"}), 404
    if not _job_access_allowed(data):
        return jsonify({"error": "المهمة تخص جلسة دخول أخرى"}), 403
    if data.get("status") in {"completed", "failed", "cancelled"}:
        return jsonify(data), 200
    updated = save_job(
        job_id,
        cancel_requested=True,
        status="cancelled",
        progress=100,
        stage="تم إيقاف التقرير",
        estimated_remaining_seconds=0,
        cancelled_at=datetime.now().isoformat(timespec="seconds"),
    )
    _release_active_report(job_id)
    return jsonify(updated), 200


def report_download_name(start_date: str, end_date: str) -> str:
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        if start.year == end.year and start.month == end.month:
            return f"Sales Report For Month {start.strftime('%B')}.xlsx"
        return f"Sales Report From {start_date} To {end_date}.xlsx"
    except Exception:
        return "Sales Report.xlsx"


@app.get("/reports/download/<job_id>")
@login_required
def download_job(job_id: str):
    data = read_json(job_path(job_id), None)
    if not data or data.get("status") != "completed":
        flash("التقرير غير جاهز للتنزيل.", "error")
        return redirect(url_for("monthly_report"))
    if not _job_access_allowed(data):
        return redirect(url_for("dashboard"))
    path = Path(data["file_path"])
    if not path.exists():
        flash("ملف التقرير غير موجود.", "error")
        return redirect(url_for("monthly_report"))
    return send_file(path, as_attachment=True, download_name=report_download_name(data["start_date"], data["end_date"]))


@app.get("/analytics")
@login_required
@analytics_required
def analytics():
    """Instant live analytics from a persistent OPOST business snapshot."""
    snapshot = load_snapshot()
    refresh_in_background(force=False)
    state = refresh_state()
    dashboard_data = read_json(DASHBOARD_DATA_PATH, {}) or {}
    summary = dashboard_data.get("summary", {}) or {}
    connection = get_database()
    identity_count = connection.execute("SELECT COUNT(*) FROM business_identities").fetchone()[0]
    connection.close()
    completed_jobs = []
    for item in JOBS_DIR.glob("*.json"):
        data = read_json(item, {}) or {}
        if data.get("status") == "completed":
            completed_jobs.append(data)
    completed_jobs.sort(key=lambda row: row.get("completed_at", ""), reverse=True)
    stats = dict(snapshot)
    stats.update({
        "identity_count": int(identity_count or 0),
        "missing_identities": max(0, int(snapshot.get("total_accounts") or 0) - int(identity_count or 0)),
        "completed_reports": len(completed_jobs),
        "best": int(summary.get("best_accounts") or 0),
        "follow_up": int(summary.get("follow_up_accounts") or 0),
        "with_shipments": int(summary.get("accounts_with_shipments") or 0),
        "without_shipments": int(summary.get("accounts_without_shipments") or 0),
        "report_period_start": dashboard_data.get("start_date", ""),
        "report_period_end": dashboard_data.get("end_date", ""),
        "report_updated_at": dashboard_data.get("updated_at", ""),
        "refreshing": state.get("refreshing", False),
    })
    return render_template("analytics.html", stats=stats, latest_jobs=completed_jobs[:6])


@app.get("/api/analytics")
@login_required
@analytics_required
def analytics_api():
    refresh_in_background(force=False)
    payload = load_snapshot()
    payload.update(refresh_state())
    return jsonify(payload)


@app.post("/api/analytics/refresh")
@login_required
@analytics_required
def analytics_refresh_api():
    try:
        payload = refresh_now()
        payload.update(refresh_state())
        return jsonify({"ok": True, **payload})
    except Exception as error:
        payload = load_snapshot()
        payload.update(refresh_state())
        return jsonify({"ok": False, "error": str(error), **payload}), 502


@app.get("/analytics/export")
@login_required
@analytics_required
def analytics_export():
    """Export a clean management-ready analytics workbook from the cached snapshot."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    snapshot = load_snapshot()
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Dashboard"
    ws.sheet_view.showGridLines = False

    navy = "24324A"
    purple = "6750A4"
    lavender = "EDE7F6"
    blue = "DCEAF7"
    mint = "DFF3E8"
    peach = "FBE7D5"
    rose = "F7DEE5"
    gold = "F5E7B8"
    soft_gray = "F5F7FA"
    border_color = "D9DFE8"
    white = "FFFFFF"
    muted = "657087"
    green = "5B8E55"
    thin = Side(style="thin", color=border_color)

    def set_fill(cell, color):
        cell.fill = PatternFill("solid", fgColor=color)

    def title(cell_range: str, text_value: str, subtitle: str = "") -> None:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = text_value
        cell.font = Font(size=22, bold=True, color=white)
        set_fill(cell, navy)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cell.row].height = 38
        if subtitle:
            row = cell.row + 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            sub = ws.cell(row, 1, subtitle)
            sub.font = Font(size=10, color=muted, italic=True)
            sub.alignment = Alignment(horizontal="left", vertical="center")

    def kpi(row: int, col: int, label: str, value, fill: str) -> None:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        label_cell = ws.cell(row, col, label)
        value_cell = ws.cell(row + 1, col, value)
        for rr in range(row, row + 3):
            for cc in range(col, col + 2):
                cell = ws.cell(rr, cc)
                set_fill(cell, fill)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        label_cell.font = Font(size=9, bold=True, color=muted)
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.font = Font(size=24, bold=True, color=navy)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

    title("A1:H1", "OPOST Analytics Report", f"Updated: {snapshot.get('updated_at', '')}  |  Source: {snapshot.get('source', 'OPOST')}")
    kpi(4, 1, "TOTAL ACCOUNTS", snapshot.get("total_accounts", 0), blue)
    kpi(4, 3, "IN INCUBATION", snapshot.get("incubation_accounts", 0), mint)
    kpi(4, 5, "ACCOUNT MANAGERS", snapshot.get("manager_count", 0), lavender)
    kpi(4, 7, "OFFICES", snapshot.get("office_count", 0), peach)

    ws["A8"] = "Monthly Growth Snapshot"
    ws["A8"].font = Font(size=14, bold=True, color=navy)
    ws["A9"], ws["B9"] = "Month", "New Accounts"
    for cell in ws[9][0:2]:
        cell.font = Font(bold=True, color=white)
        set_fill(cell, purple)
        cell.alignment = Alignment(horizontal="center")

    months = snapshot.get("months", []) or []
    recent_months = months[-12:]
    for idx, item in enumerate(recent_months, start=10):
        ws.cell(idx, 1, item.get("month"))
        ws.cell(idx, 2, item.get("count"))
        fill = soft_gray if idx % 2 == 0 else white
        for c in range(1, 3):
            set_fill(ws.cell(idx, c), fill)
            ws.cell(idx, c).border = Border(bottom=thin)
            ws.cell(idx, c).alignment = Alignment(horizontal="center")

    if recent_months:
        chart = LineChart()
        chart.title = "New accounts by month"
        chart.style = 13
        chart.y_axis.title = "Accounts"
        chart.x_axis.title = "Month"
        data = Reference(ws, min_col=2, min_row=9, max_row=9 + len(recent_months))
        cats = Reference(ws, min_col=1, min_row=10, max_row=9 + len(recent_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7.4
        chart.width = 15.2
        chart.legend = None
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showCatName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.showLeaderLines = False
        try:
            chart.series[0].graphicalProperties.line.solidFill = green
            chart.series[0].graphicalProperties.line.width = 24000
            chart.series[0].marker.symbol = "circle"
            chart.series[0].marker.size = 6
        except Exception:
            pass
        ws.add_chart(chart, "D8")

    for section_title, rows, start_col, fill in [
        ("Top Account Managers", snapshot.get("managers", [])[:8], 1, lavender),
        ("Top Offices", snapshot.get("offices", [])[:8], 5, blue),
    ]:
        row0 = 24
        ws.merge_cells(start_row=row0, start_column=start_col, end_row=row0, end_column=start_col + 1)
        heading = ws.cell(row0, start_col, section_title)
        heading.font = Font(size=13, bold=True, color=navy)
        set_fill(heading, fill)
        heading.alignment = Alignment(horizontal="center")
        ws.cell(row0 + 1, start_col, "Name")
        ws.cell(row0 + 1, start_col + 1, "Accounts")
        for c in range(start_col, start_col + 2):
            ws.cell(row0 + 1, c).font = Font(bold=True, color=white)
            set_fill(ws.cell(row0 + 1, c), purple)
            ws.cell(row0 + 1, c).alignment = Alignment(horizontal="center")
        for i, item in enumerate(rows, start=row0 + 2):
            ws.cell(i, start_col, item.get("name"))
            ws.cell(i, start_col + 1, item.get("count"))
            stripe = white if i % 2 else soft_gray
            for c in range(start_col, start_col + 2):
                set_fill(ws.cell(i, c), stripe)
                ws.cell(i, c).border = Border(bottom=thin)
            ws.cell(i, start_col + 1).alignment = Alignment(horizontal="center")

    # Dashboard widths are deliberately balanced so KPI headings never get clipped.
    for col, width in {"A":30,"B":14,"C":14,"D":14,"E":30,"F":14,"G":16,"H":16}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    ws.print_area = "A1:H34"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    def style_data_sheet(sh, headers, rows, tab_color, table_name):
        sh.sheet_view.showGridLines = False
        sh.append(headers)
        for row in rows:
            sh.append(row)
        for cell in sh[1]:
            cell.font = Font(bold=True, color=white)
            set_fill(cell, navy)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
        sh.row_dimensions[1].height = 28
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = sh.dimensions
        sh.sheet_properties.tabColor = tab_color
        if sh.max_row > 1:
            table = Table(displayName=table_name, ref=sh.dimensions)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            sh.add_table(table)
        for column in sh.columns:
            letter = column[0].column_letter
            values = [len(str(c.value or "")) for c in column[:250]]
            sh.column_dimensions[letter].width = min(44, max(12, (max(values) if values else 10) + 2))
        sh.page_setup.orientation = "landscape"
        sh.page_setup.fitToWidth = 1
        sh.sheet_properties.pageSetUpPr.fitToPage = True

    managers_rows = [[r.get("name"), r.get("count")] for r in snapshot.get("managers", [])]
    offices_rows = [[r.get("name"), r.get("count")] for r in snapshot.get("offices", [])]
    accounts_rows = [[r.get("business_id"), r.get("business_name"), r.get("created_at"), r.get("office"), r.get("account_manager"), r.get("status")] for r in snapshot.get("accounts", [])]

    sh = wb.create_sheet("Account Managers")
    style_data_sheet(sh, ["Account Manager", "Accounts"], managers_rows, "8B6FC0", "AnalyticsManagers")
    sh.column_dimensions["A"].width = 46
    sh.column_dimensions["B"].width = 14
    if managers_rows:
        max_row = min(sh.max_row, 11)  # chart only the top 10; full list remains in the table
        manager_chart = BarChart()
        manager_chart.type = "bar"
        manager_chart.style = 10
        manager_chart.title = "Top 10 account managers"
        manager_chart.y_axis.title = "Account Manager"
        manager_chart.x_axis.title = "Accounts"
        manager_chart.height = 9
        manager_chart.width = 18
        manager_chart.add_data(Reference(sh, min_col=2, min_row=1, max_row=max_row), titles_from_data=True)
        manager_chart.set_categories(Reference(sh, min_col=1, min_row=2, max_row=max_row))
        manager_chart.legend = None
        manager_chart.varyColors = True
        manager_chart.dLbls = DataLabelList()
        manager_chart.dLbls.showVal = True
        manager_chart.dLbls.showSerName = False
        manager_chart.dLbls.showCatName = False
        manager_chart.dLbls.showLegendKey = False
        sh.add_chart(manager_chart, "D2")
        sh.merge_cells("D20:L21")
        sh["D20"] = "How to read: each bar is one account manager; the number at the end of the bar is the number of businesses assigned to that manager. The complete manager list remains in the table on the left."
        sh["D20"].alignment = Alignment(wrap_text=True, vertical="top")
        sh["D20"].font = Font(size=10, color=muted, italic=True)
        set_fill(sh["D20"], lavender)

    sh = wb.create_sheet("Offices")
    style_data_sheet(sh, ["Office", "Accounts"], offices_rows, "6E9ECF", "AnalyticsOffices")
    sh.column_dimensions["A"].width = 28
    sh.column_dimensions["B"].width = 14
    office_palette = [blue, mint, lavender, peach, rose, gold]
    for row_idx in range(2, sh.max_row + 1):
        row_fill = PatternFill("solid", fgColor=office_palette[(row_idx - 2) % len(office_palette)])
        for col_idx in range(1, min(sh.max_column, 2) + 1):
            sh.cell(row_idx, col_idx).fill = row_fill
            sh.cell(row_idx, col_idx).font = Font(color=navy, bold=(col_idx == 1))
            sh.cell(row_idx, col_idx).alignment = Alignment(horizontal="center" if col_idx == 2 else "left")
    if offices_rows:
        max_row = min(sh.max_row, 11)
        office_chart = BarChart()
        office_chart.type = "bar"
        office_chart.style = 10
        office_chart.title = "Top 10 offices"
        office_chart.y_axis.title = "Office"
        office_chart.x_axis.title = "Accounts"
        office_chart.height = 9
        office_chart.width = 18
        office_chart.add_data(Reference(sh, min_col=2, min_row=1, max_row=max_row), titles_from_data=True)
        office_chart.set_categories(Reference(sh, min_col=1, min_row=2, max_row=max_row))
        office_chart.legend = None
        office_chart.varyColors = True
        office_chart.dLbls = DataLabelList()
        office_chart.dLbls.showVal = True
        office_chart.dLbls.showSerName = False
        office_chart.dLbls.showCatName = False
        office_chart.dLbls.showLegendKey = False
        sh.add_chart(office_chart, "D2")
        sh.merge_cells("D20:L21")
        sh["D20"] = "How to read: every color identifies a different office in the table. The chart shows the 10 offices with the most businesses; the full office list remains available on the left."
        sh["D20"].alignment = Alignment(wrap_text=True, vertical="top")
        sh["D20"].font = Font(size=10, color=muted, italic=True)
        set_fill(sh["D20"], blue)

    # Full monthly history stays in the table; the chart intentionally focuses on the latest 24 months.
    sh = wb.create_sheet("Monthly Growth")
    sh.sheet_view.showGridLines = False
    sh.sheet_properties.tabColor = "D5A94E"
    sh.merge_cells("A1:F1")
    sh["A1"] = "Monthly Account Growth"
    sh["A1"].font = Font(size=20, bold=True, color=white)
    set_fill(sh["A1"], navy)
    sh["A1"].alignment = Alignment(horizontal="center")
    sh.merge_cells("A2:F2")
    sh["A2"] = "Full month-by-month history of new OPOST accounts"
    sh["A2"].font = Font(size=10, italic=True, color=muted)
    sh["A2"].alignment = Alignment(horizontal="center")
    sh["A4"], sh["B4"] = "Month", "New Accounts"
    for cell in sh[4][0:2]:
        cell.font = Font(bold=True, color=white)
        set_fill(cell, purple)
        cell.alignment = Alignment(horizontal="center")
    for i, item in enumerate(months, start=5):
        sh.cell(i, 1, item.get("month"))
        sh.cell(i, 2, item.get("count"))
        row_fill = office_palette[(i - 5) % len(office_palette)]
        for c in range(1, 3):
            set_fill(sh.cell(i, c), row_fill)
            sh.cell(i, c).border = Border(bottom=thin)
            sh.cell(i, c).alignment = Alignment(horizontal="center")
    sh.column_dimensions["A"].width = 18
    sh.column_dimensions["B"].width = 18
    sh.column_dimensions["C"].width = 3
    sh.freeze_panes = "A5"

    if months:
        chart_months = months[-24:]
        chart_start_data_row = 5 + max(0, len(months) - len(chart_months))
        chart_end_row = 4 + len(months)
        growth_chart = LineChart()
        growth_chart.title = "Monthly growth trend — latest 24 months"
        growth_chart.style = 13
        growth_chart.y_axis.title = "New Accounts"
        growth_chart.x_axis.title = "Month"
        growth_chart.add_data(Reference(sh, min_col=2, min_row=chart_start_data_row, max_row=chart_end_row), titles_from_data=False)
        growth_chart.set_categories(Reference(sh, min_col=1, min_row=chart_start_data_row, max_row=chart_end_row))
        growth_chart.height = 9
        growth_chart.width = 19
        growth_chart.legend = None
        try:
            growth_chart.series[0].graphicalProperties.line.solidFill = green
            growth_chart.series[0].graphicalProperties.line.width = 26000
            growth_chart.series[0].marker.symbol = "circle"
            growth_chart.series[0].marker.size = 5
        except Exception:
            pass
        sh.add_chart(growth_chart, "D4")

        counts = [int(item.get("count") or 0) for item in chart_months]
        latest = chart_months[-1]
        peak = max(chart_months, key=lambda item: int(item.get("count") or 0))
        average = round(sum(counts) / len(counts), 1) if counts else 0
        sh.merge_cells("D22:L25")
        sh["D22"] = (
            f"Simple reading: the line shows how many new accounts were created each month. "
            f"Latest month: {latest.get('month')} = {latest.get('count')} accounts. "
            f"Highest month in this 24-month view: {peak.get('month')} = {peak.get('count')} accounts. "
            f"Average: {average} accounts/month. The full historical data is kept in the table on the left."
        )
        sh["D22"].alignment = Alignment(wrap_text=True, vertical="top")
        sh["D22"].font = Font(size=10, color=navy)
        set_fill(sh["D22"], gold)
        sh["D22"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.sheet_properties.pageSetUpPr.fitToPage = True

    sh = wb.create_sheet("All Accounts")
    style_data_sheet(sh, ["Business ID", "Business Name", "Created At", "Office", "Account Manager", "Status"], accounts_rows, "7DAA91", "AnalyticsAccounts")
    sh.column_dimensions["A"].width = 16
    sh.column_dimensions["B"].width = 34
    sh.column_dimensions["C"].width = 22
    sh.column_dimensions["D"].width = 22
    sh.column_dimensions["E"].width = 38
    sh.column_dimensions["F"].width = 14
    for row_idx in range(2, sh.max_row + 1):
        sh.cell(row_idx, 1).alignment = Alignment(horizontal="center")
        sh.cell(row_idx, 3).alignment = Alignment(horizontal="center")
        sh.cell(row_idx, 6).alignment = Alignment(horizontal="center")

    output = DOWNLOADS_DIR / f"OPOST Analytics {datetime.now().strftime('%Y-%m-%d %H-%M')}.xlsx"
    wb.save(output)
    return send_file(output, as_attachment=True, download_name=output.name)


@app.get("/download-center")
@login_required
def download_center():
    """List generated Excel and PDF files without contacting OPOST."""
    files = []
    seen = set()
    # Current persistent archive.
    for path in DOWNLOADS_DIR.glob("*"):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".pdf"}:
            continue
        resolved = str(path.resolve())
        seen.add(resolved)
        stat = path.stat()
        files.append({
            "name": path.name,
            "type": "Excel" if path.suffix.lower() == ".xlsx" else "PDF",
            "size": _format_file_size(stat.st_size),
            "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "timestamp": stat.st_mtime,
            "download_url": url_for("download_center_file", filename=path.name),
        })
    # Preserve access to old completed reports created before 3.0.0.
    for item in JOBS_DIR.glob("*.json"):
        data = read_json(item, {}) or {}
        if data.get("status") != "completed" or not data.get("file_path"):
            continue
        source = Path(data["file_path"])
        if not source.exists() or source.suffix.lower() != ".xlsx":
            continue
        resolved = str(source.resolve())
        if resolved in seen:
            continue
        try:
            archived = _archive_generated_file(source, report_download_name(data.get("start_date", ""), data.get("end_date", "")))
            seen.add(str(archived.resolve()))
            stat = archived.stat()
            files.append({
                "name": archived.name, "type": "Excel", "size": _format_file_size(stat.st_size),
                "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "timestamp": stat.st_mtime, "download_url": url_for("download_center_file", filename=archived.name),
            })
        except Exception:
            pass
    files.sort(key=lambda row: row["timestamp"], reverse=True)
    return render_template("download_center.html", files=files)


@app.get("/download-center/file/<path:filename>")
@login_required
def download_center_file(filename: str):
    safe_name = Path(filename).name
    path = (DOWNLOADS_DIR / safe_name).resolve()
    try:
        path.relative_to(DOWNLOADS_DIR.resolve())
    except ValueError:
        abort(404)
    if not path.exists() or path.suffix.lower() not in {".xlsx", ".pdf"}:
        abort(404)
    return send_file(path, as_attachment=True, download_name=path.name, max_age=0)


@app.route("/follow-up")
@login_required
def follow_up():
    return render_template("follow_up.html", follow_up_accounts=read_json(FOLLOW_UP_DATA_PATH, []))


@app.route("/account-manager-transfer", methods=["GET", "POST"])
@account_manager_transfer_required
def account_manager_transfer():
    """Render instantly; OPOST data is loaded only after the user requests it."""
    transfer_result = None
    source_manager_id = str(request.form.get("source_manager_id") or "").strip()
    target_manager_id = str(request.form.get("target_manager_id") or "").strip()
    start_date = str(request.form.get("start_date") or "").strip()
    end_date = str(request.form.get("end_date") or "").strip()

    cached = read_json(ACCOUNT_MANAGERS_CACHE_PATH, {}) or {}
    managers = cached.get("items", []) if isinstance(cached, dict) else []
    manager_by_id = {str(item.get("id")): str(item.get("name") or "") for item in managers if isinstance(item, dict)}
    source_manager_label = manager_by_id.get(source_manager_id, "")
    target_manager_label = manager_by_id.get(target_manager_id, "")

    if request.method == "POST" and request.form.get("action") == "transfer":
        selected_ids = request.form.getlist("business_ids")
        if not selected_ids:
            flash("اختر حسابًا واحدًا على الأقل للتحويل.", "error")
        elif not source_manager_id or not target_manager_id:
            flash("اختر المدير الحالي والمدير الجديد.", "error")
        else:
            try:
                transfer_result = opost_bulk_change_account_manager(
                    selected_ids,
                    source_manager_id,
                    source_manager_label,
                    target_manager_id,
                    target_manager_label,
                    max_workers=6,
                )
                try:
                    posted_accounts = json.loads(request.form.get("selected_accounts_json") or "[]")
                except Exception:
                    posted_accounts = []
                posted_by_id = {str(item.get("id") or ""): item for item in posted_accounts if isinstance(item, dict)}
                for bucket in ("success", "failed"):
                    for item in transfer_result.get(bucket, []):
                        meta = posted_by_id.get(str(item.get("id") or ""), {})
                        item["name"] = str(meta.get("name") or "")
                        item["office"] = str(meta.get("office") or "")
                        item["created_at"] = str(meta.get("created_at") or "")
                success_count = len(transfer_result.get("success", []))
                failed_count = len(transfer_result.get("failed", []))
                if success_count:
                    flash(f"تم تحويل {success_count} حساب فعليًا داخل OPOST.", "success")
                if failed_count:
                    flash(f"تعذر تحويل {failed_count} حساب. راجع تفاصيل النتائج.", "error")
            except Exception as error:
                flash(f"تعذر تنفيذ التحويل داخل OPOST: {error}", "error")

    return render_template(
        "account_manager_transfer.html",
        start_date=start_date,
        end_date=end_date,
        transfer_result=transfer_result,
        managers=managers,
        source_manager_id=source_manager_id,
        target_manager_id=target_manager_id,
        source_manager=source_manager_label,
        target_manager=target_manager_label,
    )


def _load_account_managers(force: bool = False) -> list[dict[str, str]]:
    now = time.time()
    with _ACCOUNT_MANAGERS_LOCK:
        cached = read_json(ACCOUNT_MANAGERS_CACHE_PATH, {}) or {}
        if not force and isinstance(cached, dict):
            # A previously verified manager list is safe to show immediately.
            # Do not block the page merely because its refresh timestamp expired;
            # the user can request a manual refresh when needed.
            items = cached.get("items") or []
            if cached.get("schema_version") == 4 and items:
                return items
        client = OpostClient()
        try:
            client.start()
            client.login()
            items = opost_list_account_managers(client)
        finally:
            client.close()
        write_json(ACCOUNT_MANAGERS_CACHE_PATH, {"schema_version": 4, "saved_at": now, "items": items})
        return items


@app.get("/api/account-managers")
@account_manager_transfer_required
def api_account_managers():
    force = request.args.get("refresh") == "1"
    try:
        items = _load_account_managers(force=force)
        return jsonify({"ok": True, "items": items})
    except Exception as error:
        cached = read_json(ACCOUNT_MANAGERS_CACHE_PATH, {}) or {}
        items = cached.get("items", []) if isinstance(cached, dict) else []
        return jsonify({"ok": bool(items), "items": items, "error": str(error)}), (200 if items else 503)


@app.get("/api/account-manager-candidates")
@account_manager_transfer_required
def api_account_manager_candidates():
    start_date = str(request.args.get("start_date") or "").strip()
    end_date = str(request.args.get("end_date") or "").strip()
    source_manager_id = str(request.args.get("source_manager_id") or "").strip()
    source_manager_label = str(request.args.get("source_manager_label") or "").strip()
    if not start_date or not end_date or not source_manager_id:
        return jsonify({"ok": False, "error": "missing_filters"}), 400
    try:
        client = OpostClient()
        try:
            client.start()
            client.login()
            items = opost_find_manager_candidates(
                client,
                start_date,
                end_date,
                source_manager_id,
                source_manager_label,
            )
        finally:
            client.close()
        return jsonify({"ok": True, "items": items, "count": len(items)})
    except Exception as error:
        return jsonify({"ok": False, "error": str(error)}), 502


@app.route("/employees", methods=["GET", "POST"])
@users_required
def employees():
    connection = get_database()
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role", "employee")
        can_ids = 1 if request.form.get("can_manage_identities") == "1" else 0
        can_transfer = 1 if request.form.get("can_transfer_account_managers") == "1" else 0
        can_analytics = 1 if request.form.get("can_view_analytics") == "1" else 0
        can_users = 1 if request.form.get("can_manage_users") == "1" else 0
        if not full_name or not email or len(password) < 6 or role not in {"admin", "employee"}:
            flash("تحقق من الاسم والبريد وكلمة المرور والصلاحية.", "error")
        else:
            try:
                connection.execute(
                    "INSERT INTO employees(full_name,email,password_hash,role,active,can_manage_identities,can_transfer_account_managers,can_view_analytics,can_manage_users,created_at) VALUES(?,?,?,?,1,?,?,?,?,?)",
                    (full_name, email, generate_password_hash(password), role, can_ids, can_transfer, can_analytics, can_users, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                )
                connection.commit()
                backup_users()
                flash("تمت إضافة المستخدم.", "success")
            except sqlite3.IntegrityError:
                flash("البريد مستخدم مسبقًا.", "error")
    rows = connection.execute("SELECT id,full_name,email,role,active,can_manage_identities,can_transfer_account_managers,can_view_analytics,can_manage_users,created_at FROM employees ORDER BY id DESC").fetchall()
    active_users = connection.execute("SELECT COUNT(*) FROM employees WHERE active=1").fetchone()[0]
    total_users = connection.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    connection.close()
    return render_template("employees.html", employees=rows, active_users=int(active_users or 0), total_users=int(total_users or 0))


@app.route("/employees/<int:employee_id>/edit", methods=["POST"])
@users_required
def edit_employee(employee_id: int):
    full_name = request.form.get("full_name", "").strip()
    email = request.form.get("email", "").strip()
    role = request.form.get("role", "employee").strip()
    active = 1 if request.form.get("active") == "1" else 0
    can_ids = 1 if request.form.get("can_manage_identities") == "1" else 0
    can_transfer = 1 if request.form.get("can_transfer_account_managers") == "1" else 0
    can_analytics = 1 if request.form.get("can_view_analytics") == "1" else 0
    can_users = 1 if request.form.get("can_manage_users") == "1" else 0
    password = request.form.get("password", "")

    if not full_name or not email or role not in {"admin", "employee"}:
        flash("تحقق من الاسم والبريد والصلاحية.", "error")
        return redirect(url_for("employees"))
    if password and len(password) < 6:
        flash("كلمة المرور الجديدة يجب ألا تقل عن 6 أحرف.", "error")
        return redirect(url_for("employees"))
    if employee_id == session.get("employee_id") and active == 0:
        flash("لا يمكنك تعطيل حسابك الحالي.", "error")
        return redirect(url_for("employees"))

    connection = get_database()
    try:
        values = [full_name, email, role, active, can_ids, can_transfer, can_analytics, can_users]
        sql = "UPDATE employees SET full_name=?, email=?, role=?, active=?, can_manage_identities=?, can_transfer_account_managers=?, can_view_analytics=?, can_manage_users=?"
        if password:
            sql += ", password_hash=?"
            values.append(generate_password_hash(password))
        sql += " WHERE id=?"
        values.append(employee_id)
        connection.execute(sql, tuple(values))
        connection.commit()
        backup_users()
        if employee_id == session.get("employee_id"):
            session["employee_name"] = full_name
            session["employee_role"] = role
        flash("تم تحديث الحساب.", "success")
    except sqlite3.IntegrityError:
        flash("البريد الإلكتروني مستخدم مسبقًا.", "error")
    finally:
        connection.close()
    return redirect(url_for("employees"))


@app.post("/employees/<int:employee_id>/toggle")
@users_required
def toggle_employee(employee_id: int):
    if employee_id == session.get("employee_id"):
        flash("لا يمكنك تعطيل حسابك الحالي.", "error")
    else:
        connection = get_database()
        connection.execute("UPDATE employees SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=?", (employee_id,))
        connection.commit()
        connection.close()
        backup_users()
    return redirect(url_for("employees"))


@app.post("/employees/<int:employee_id>/delete")
@users_required
def delete_employee(employee_id: int):
    """Delete a site-login account for admins/users with user-management permission.

    The currently signed-in account and the built-in recovery admin are protected
    to avoid accidental lockout or an account that would be recreated on restart.
    """
    try:
        current_id = int(session.get("employee_id"))
    except (TypeError, ValueError):
        current_id = 0

    if employee_id == current_id:
        flash("لا يمكنك حذف الحساب الذي تستخدمه حاليًا.", "error")
        return redirect(url_for("employees"))

    connection = get_database()
    try:
        employee = connection.execute(
            "SELECT id,email FROM employees WHERE id=?", (employee_id,)
        ).fetchone()
        if not employee:
            flash("الحساب المطلوب غير موجود.", "error")
            return redirect(url_for("employees"))

        # This account is the local recovery administrator requested as the
        # permanent baseline account, so removing it would only make it return
        # on the next start and could be confusing to administrators.
        if str(employee["email"] or "") == "Mansour_E@gmail.com":
            flash("الحساب الإداري الأساسي محمي من الحذف.", "error")
            return redirect(url_for("employees"))

        connection.execute("DELETE FROM employees WHERE id=?", (employee_id,))
        connection.commit()
        backup_users()
        flash("تم حذف المستخدم من حسابات الدخول للموقع.", "success")
    finally:
        connection.close()
    return redirect(url_for("employees"))



def _identity_source_folder() -> str:
    try:
        payload = json.loads(IDENTITY_SETTINGS_PATH.read_text(encoding="utf-8"))
        return str(payload.get("source_folder") or "")
    except Exception:
        return ""

def _save_identity_source_folder(folder: str) -> None:
    USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    write_json(IDENTITY_SETTINGS_PATH, {"source_folder": folder})

_IDENTITY_INDEX_CACHE = None
_IDENTITY_INDEX_LOCK = threading.RLock()

def _business_identity_index():
    global _IDENTITY_INDEX_CACHE
    with _IDENTITY_INDEX_LOCK:
        if _IDENTITY_INDEX_CACHE is not None:
            return _IDENTITY_INDEX_CACHE
    from web_app.business_profile_service import get_business_name_index, DIRECTORY_CACHE
    from opost_client import OpostClient
    client = OpostClient()
    try:
        client.start()
        # Identity matching is local after the first directory warm-up. Avoid an
        # OPOST authentication probe on every 4–10 image update.
        if not DIRECTORY_CACHE.exists():
            client.login()
        raw_index = get_business_name_index(client)
    finally:
        client.close()
    index = {}
    for normalized, items in raw_index.items():
        for item in items:
            name = str(item.get("display") or item.get("name") or field_value(item, "name") or "").strip()
            if name:
                index.setdefault(normalized, []).append((str(item.get("id")), name))
    with _IDENTITY_INDEX_LOCK:
        _IDENTITY_INDEX_CACHE = index
    return index

def _identity_sync_status_path(employee_id: int) -> Path:
    return USER_DATA_DIR / f"identity_sync_status_{int(employee_id)}.json"


def _write_identity_sync_status(employee_id: int, **updates: Any) -> dict[str, Any]:
    path = _identity_sync_status_path(employee_id)
    current = read_json(path, {}) or {}
    current.update(updates)
    current["updated_at"] = datetime.now().isoformat(timespec="seconds")
    write_json(path, current)
    return current


def _sync_identity_folder(folder: Path, uploaded_by: int):
    """Synchronize identities without depending on the browser request.

    The work runs in a background executor, so switching tabs, opening Excel,
    or navigating elsewhere cannot cancel the server-side synchronization.
    Existing SQLite rows are prefetched once to avoid two SELECT queries for
    every image when folders contain hundreds or thousands of files.
    """
    import hashlib
    started = time.perf_counter()
    index = _business_identity_index()
    connection = get_database()
    results = []
    scanned = changed = 0
    try:
        previous_rows = connection.execute("SELECT * FROM identity_sync_files").fetchall()
        previous_by_source = {str(row["source_path"]): row for row in previous_rows}
        identity_rows = connection.execute("SELECT * FROM business_identities").fetchall()
        identity_by_business = {str(row["business_id"]): row for row in identity_rows}

        for entry in folder.rglob("*"):
            if not entry.is_file() or entry.suffix.lower() not in ALLOWED_IDENTITY_EXTENSIONS:
                continue
            scanned += 1
            stat = entry.stat()
            source = str(entry.resolve())
            previous = previous_by_source.get(source)
            if previous and float(previous["file_mtime"]) == float(stat.st_mtime) and int(previous["file_size"]) == int(stat.st_size):
                continue
            changed += 1
            matches = index.get(normalize_business_name(entry.stem), [])
            if len(matches) != 1:
                msg = "لم يتم العثور على حساب مطابق" if not matches else "يوجد أكثر من حساب بنفس الاسم"
                results.append({"file": entry.name, "status": "warning", "message": msg})
                continue
            business_id, business_name = matches[0]
            content = entry.read_bytes()
            digest = hashlib.sha256(content).hexdigest()
            existing = identity_by_business.get(str(business_id))
            folder_dest = IDENTITY_DIR / business_id
            folder_dest.mkdir(parents=True, exist_ok=True)
            stored = f"identity{entry.suffix.lower()}"
            destination = folder_dest / stored
            if not existing or existing["file_hash"] != digest:
                destination.write_bytes(content)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            connection.execute(
                """INSERT INTO business_identities(business_id,business_name,stored_filename,original_filename,file_hash,uploaded_by,uploaded_at,updated_at)
                   VALUES(?,?,?,?,?,?,?,?)
                   ON CONFLICT(business_id) DO UPDATE SET business_name=excluded.business_name,stored_filename=excluded.stored_filename,
                   original_filename=excluded.original_filename,file_hash=excluded.file_hash,uploaded_by=excluded.uploaded_by,updated_at=excluded.updated_at""",
                (business_id,business_name,stored,entry.name,digest,int(uploaded_by),now,now),
            )
            connection.execute(
                """INSERT INTO identity_sync_files(source_path,file_mtime,file_size,file_hash,business_id,synced_at)
                   VALUES(?,?,?,?,?,?)
                   ON CONFLICT(source_path) DO UPDATE SET file_mtime=excluded.file_mtime,file_size=excluded.file_size,
                   file_hash=excluded.file_hash,business_id=excluded.business_id,synced_at=excluded.synced_at""",
                (source,stat.st_mtime,stat.st_size,digest,business_id,now),
            )
            results.append({"file": entry.name, "status": "success", "message": "تمت المزامنة", "business": business_name, "business_id": business_id})
        connection.commit()
    finally:
        connection.close()
    return results, scanned, changed, round(time.perf_counter()-started, 2)


def _run_identity_sync_background(folder: Path, employee_id: int) -> None:
    global _IDENTITY_SYNC_ACTIVE
    try:
        _write_identity_sync_status(employee_id, status="running", scanned=0, changed=0, elapsed=0, results=[])
        results, scanned, changed, elapsed = _sync_identity_folder(folder, employee_id)
        review = [item for item in results if item.get("status") != "success"][-50:]
        if not review:
            review = results[-20:]
        _write_identity_sync_status(
            employee_id, status="completed", scanned=scanned, changed=changed,
            elapsed=elapsed, results=review, error="",
        )
    except Exception as error:
        _write_identity_sync_status(employee_id, status="failed", error=str(error))
    finally:
        with _IDENTITY_SYNC_LOCK:
            _IDENTITY_SYNC_ACTIVE = False

@app.post("/identities/select-folder")
@identities_required
def select_identity_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        folder = filedialog.askdirectory(title="اختر مجلد هويات الحسابات")
        root.destroy()
        if folder:
            _save_identity_source_folder(folder)
            flash("تم حفظ مجلد الهويات.", "success")
    except Exception as error:
        flash(f"تعذر فتح اختيار المجلد: {error}", "error")
    return redirect(url_for("identities"))

@app.post("/identities/sync-folder")
@identities_required
def sync_identity_folder():
    global _IDENTITY_SYNC_ACTIVE
    folder_text = _identity_source_folder()
    folder = Path(folder_text) if folder_text else None
    if not folder or not folder.exists() or not folder.is_dir():
        flash("اختر مجلد الهويات أولًا.", "error")
        return redirect(url_for("identities"))
    employee_id = int(session["employee_id"])
    with _IDENTITY_SYNC_LOCK:
        if _IDENTITY_SYNC_ACTIVE:
            flash("مزامنة الهويات تعمل حاليًا في الخلفية.", "info")
            return redirect(url_for("identities"))
        _IDENTITY_SYNC_ACTIVE = True
    _write_identity_sync_status(employee_id, status="queued", results=[], error="")
    _IDENTITY_SYNC_EXECUTOR.submit(_run_identity_sync_background, folder, employee_id)
    flash("بدأت مزامنة الهويات في الخلفية. يمكنك التنقل أو فتح Excel ولن تتوقف العملية.", "success")
    return redirect(url_for("identities"))


@app.get("/api/identities/sync-status")
@identities_required
def identity_sync_status_api():
    employee_id = int(session["employee_id"])
    return jsonify(read_json(_identity_sync_status_path(employee_id), {"status": "idle"}) or {"status": "idle"})


@app.get("/identities")
@identities_required
def identities():
    employee_id = int(session["employee_id"])
    sync_status = read_json(_identity_sync_status_path(employee_id), {}) or {}
    results = sync_status.get("results", []) if sync_status.get("status") in {"completed", "failed"} else []
    connection = get_database()
    records = connection.execute(
        "SELECT bi.*, e.full_name uploaded_by_name FROM business_identities bi "
        "LEFT JOIN employees e ON e.id=bi.uploaded_by ORDER BY bi.updated_at DESC"
    ).fetchall()
    connection.close()
    return render_template(
        "identities.html",
        records=records,
        results=results,
        source_folder=_identity_source_folder(),
        identity_sync_status=sync_status,
    )


@app.get("/identities/image/<business_id>")
@login_required
def identity_image(business_id: str):
    if not can_manage_identities():
        abort(403)
    row = get_identity_record(business_id)
    if not row:
        abort(404)
    path = IDENTITY_DIR / str(business_id) / row["stored_filename"]
    if not path.exists():
        abort(404)
    return send_file(path, conditional=True)


@app.get("/health")
def health():
    return jsonify({"status": "ok", "version": "3.0.15"})



def _warm_runtime_caches() -> None:
    """Warm only missing caches, and never compete with the first page load.

    Previous builds started a full business-directory fetch, manager discovery,
    and analytics refresh almost at the same time.  Those jobs all hit OPOST and
    made an otherwise local dashboard feel slow.  Existing caches are now used
    immediately; network warm-up happens only when a required cache is missing.
    """
    # Let Flask become responsive before any optional OPOST work starts.
    time.sleep(2.0)
    try:
        from web_app.business_profile_service import DIRECTORY_CACHE
        # The directory is already persistent and valid for fast search.  Do not
        # log in to OPOST at startup merely to read a file we already have.
        if not DIRECTORY_CACHE.exists():
            from opost_client import OpostClient
            from web_app.business_profile_service import _load_directory
            client = OpostClient()
            try:
                client.start(); client.login(); _load_directory(client)
            finally:
                client.close()

        # Likewise, only discover managers when no verified local list exists.
        cached_managers = read_json(ACCOUNT_MANAGERS_CACHE_PATH, {}) or {}
        if not (isinstance(cached_managers, dict) and cached_managers.get("schema_version") == 4 and cached_managers.get("items")):
            try:
                _load_account_managers(force=False)
            except Exception as manager_error:
                print(f"Account-manager warm-up skipped: {manager_error}")
    except Exception as error:
        print(f"Cache warm-up skipped: {error}")


migrate_legacy_database()
initialize_database()
create_default_admin()
threading.Thread(target=_warm_runtime_caches, name="opost-cache-warm", daemon=True).start()

# Do not launch a full Analytics refresh at application startup.  The Analytics
# page already refreshes itself in the background when it is actually opened.
# This keeps login/dashboard/navigation responsive and avoids duplicate OPOST
# pagination competing with report, search, or identity jobs.

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
