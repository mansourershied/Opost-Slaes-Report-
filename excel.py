from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ==========================================
# Colors
# ==========================================

# Elegant light report palette (independent from website colors)
BLUE = "6F7F9B"          # muted steel blue for sheet titles
LIGHT_BLUE = "E4EBF3"    # soft blue-grey for regular headers
GREEN = "DDEBDD"         # soft sage for positive performance
GREEN_ROW = "EEF6EE"
RED = "F2DCDC"           # muted blush for negative performance
RED_ROW = "FAEEEE"
LAVENDER = "E9E1F0"      # soft lavender for secondary metrics
PEACH = "F3E5D8"         # soft peach for returns/follow-up
YELLOW = "F3EBCF"        # restrained cream accent
ORANGE = "F1E1D5"
WHITE = "FFFFFF"
DARK_TEXT = "26354A"
MUTED_TEXT = "5E6878"
GRAY = "F7F9FC"


THIN = Side(
    style="thin",
    color="D9D9D9"
)

BORDER = Border(
    left=THIN,
    right=THIN,
    top=THIN,
    bottom=THIN
)


# ==========================================
# Auto Width
# ==========================================

def auto_width(ws):

    for column_number in range(
        1,
        ws.max_column + 1
    ):

        column_letter = get_column_letter(
            column_number
        )

        maximum_length = 0

        for row_number in range(
            1,
            ws.max_row + 1
        ):

            cell = ws.cell(
                row=row_number,
                column=column_number
            )

            if cell.value is None:
                continue

            maximum_length = max(
                maximum_length,
                len(str(cell.value))
            )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 3, 12),
            35
        )


# ==========================================
# Title
# ==========================================

def create_title(
    ws,
    title_text,
    number_of_columns
):

    number_of_columns = max(
        number_of_columns,
        1
    )

    last_column = get_column_letter(
        number_of_columns
    )

    ws.merge_cells(
        f"A1:{last_column}1"
    )

    title = ws["A1"]

    title.value = title_text

    title.font = Font(
        bold=True,
        size=18,
        color=WHITE
    )

    title.fill = PatternFill(
        "solid",
        fgColor=BLUE
    )

    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30


# ==========================================
# Empty Sheet
# ==========================================

def create_empty_sheet(
    ws,
    title_text,
    message
):

    create_title(
        ws,
        title_text,
        6
    )

    ws["A3"] = message

    ws["A3"].font = Font(
        italic=True,
        color="666666"
    )

    ws.column_dimensions["A"].width = 50


# ==========================================
# Export Table
# ==========================================

def export_sheet(
    ws,
    rows,
    title_text,
    empty_message="No Data"
):

    # Keep internal status counts available to the calculations, but never
    # expose the technical/unknown-status column in the Excel deliverable.
    excluded_columns = {"Unknown Status", "In Progress / Other"}
    rows = [
        {key: value for key, value in row.items() if key not in excluded_columns}
        for row in (rows or [])
        if isinstance(row, dict)
    ]

    if not rows:

        create_empty_sheet(
            ws,
            title_text,
            empty_message
        )

        return

    headers = list(
        rows[0].keys()
    )

    create_title(
        ws,
        title_text,
        len(headers)
    )


    # ======================================
    # Headers
    # ======================================

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=3,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color=DARK_TEXT
        )

        header_color = LIGHT_BLUE

        if header == "Closed %":
            header_color = GREEN
        elif header == "Delivered %":
            header_color = RED
        elif header == "Returned %":
            header_color = PEACH
        elif header == "Cancelled %":
            header_color = LAVENDER

        cell.fill = PatternFill(
            "solid",
            fgColor=header_color
        )

        cell.border = BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


    # ======================================
    # Rows
    # ======================================

    for row_number, row_data in enumerate(
        rows,
        start=4
    ):

        for column_number, header in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=row_number,
                column=column_number
            )

            cell.value = row_data.get(
                header,
                ""
            )

            cell.border = BORDER

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            # Light, restrained metric fills
            if header == "Closed %":
                cell.fill = PatternFill("solid", fgColor=GREEN_ROW)
                cell.number_format = '0.00"%"'
            elif header == "Delivered %":
                cell.fill = PatternFill("solid", fgColor=RED_ROW)
                cell.number_format = '0.00"%"'
            elif header == "Returned %":
                cell.fill = PatternFill("solid", fgColor="FBF3EC")
                cell.number_format = '0.00"%"'
            elif header == "Cancelled %":
                cell.fill = PatternFill("solid", fgColor="F5F0F8")
                cell.number_format = '0.00"%"'
            elif row_number % 2 == 0:

                cell.fill = PatternFill(
                    "solid",
                    fgColor=GRAY
                )


    ws.freeze_panes = "A4"

    ws.auto_filter.ref = (
        f"A3:"
        f"{get_column_letter(len(headers))}"
        f"{ws.max_row}"
    )

    ws.sheet_view.showGridLines = False

    auto_width(ws)


# ==========================================
# Ranking
# ==========================================

def add_ranking(rows):

    ranked_rows = []

    for rank, row_data in enumerate(
        rows,
        start=1
    ):

        ranked_row = {
            "Rank": rank
        }

        ranked_row.update(
            row_data
        )

        ranked_rows.append(
            ranked_row
        )

    return ranked_rows


# ==========================================
# Summary
# ==========================================

def export_summary(
    ws,
    summary
):

    create_title(
        ws,
        "OPOST MONTHLY INCUBATION REPORT",
        4
    )


    # ======================================
    # Summary Headers
    # ======================================

    headers = [
        "Metric",
        "Count",
        "Percentage",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=3,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True,
            color=DARK_TEXT
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=LIGHT_BLUE
        )

        cell.border = BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # ======================================
    # Summary Rows
    # ======================================

    summary_rows = [
        (
            "Accounts Created",
            summary.get(
                "total_accounts",
                0
            ),
            100.0,
            LIGHT_BLUE
        ),

        (
            "Accounts With Shipments",
            summary.get(
                "accounts_with_shipments",
                0
            ),
            summary.get(
                "accounts_with_shipments_percentage",
                0
            ),
            GREEN
        ),

        (
            "Accounts Without Shipments",
            summary.get(
                "accounts_without_shipments",
                0
            ),
            summary.get(
                "accounts_without_shipments_percentage",
                0
            ),
            RED
        ),

        (
            "Best Accounts",
            summary.get(
                "best_accounts",
                0
            ),
            None,
            GREEN
        ),

        (
            "Need Follow Up",
            summary.get(
                "follow_up_accounts",
                0
            ),
            None,
            ORANGE
        ),
    ]


    start_row = 4

    for index, summary_row in enumerate(
        summary_rows
    ):

        row_number = (
            start_row + index
        )

        label = summary_row[0]
        value = summary_row[1]
        percentage = summary_row[2]
        color = summary_row[3]

        label_cell = ws.cell(
            row=row_number,
            column=1
        )

        value_cell = ws.cell(
            row=row_number,
            column=2
        )

        percentage_cell = ws.cell(
            row=row_number,
            column=3
        )

        label_cell.value = label

        value_cell.value = value

        if percentage is not None:

            percentage_cell.value = (
                float(percentage) / 100
            )

            percentage_cell.number_format = (
                "0.00%"
            )

        else:

            percentage_cell.value = "-"

        label_cell.font = Font(
            bold=True
        )

        for cell in (
            label_cell,
            value_cell,
            percentage_cell
        ):

            cell.fill = PatternFill(
                "solid",
                fgColor=color
            )

            cell.border = BORDER

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )


    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ws.freeze_panes = "A4"

    ws.sheet_view.showGridLines = False



REPORT_COLUMN_ORDER = [
    "Business ID", "Business Name", "Created At", "Account Age",
    "Shipments", "Delivered", "Successful Delivery %",
    "Returned", "Returned %", "Status", "Category", "Follow Up Reason",
]

def order_report_rows(rows):
    ordered=[]
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        item={key: row.get(key, "") for key in REPORT_COLUMN_ORDER}
        # Preserve a few operational fields after the requested report columns.
        for key in ("Phone", "Account Manager", "Office", "Closed", "Cancelled"):
            if key in row:
                item[key]=row.get(key, "")
        ordered.append(item)
    return ordered

# ==========================================
# Main Export
# ==========================================

def export_businesses(
    summary,
    all_accounts,
    best_accounts,
    follow_up_accounts,
    no_shipments,
    start_date=None,
    end_date=None
):

    workbook = Workbook()

    # The complete account list is the primary report.
    # Keep it as the first and active sheet so the user immediately
    # sees every account returned by OPOST, not only the ranked subset.
    all_accounts_sheet = workbook.active
    all_accounts_sheet.title = "All Businesses"

    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    best_sheet = workbook.create_sheet(
        "Best Accounts"
    )

    follow_up_sheet = workbook.create_sheet(
        "Need Follow Up"
    )

    no_shipments_sheet = workbook.create_sheet(
        "No Shipments"
    )

    # Soft, distinct tab colors for quick navigation.
    all_accounts_sheet.sheet_properties.tabColor = "B7C7D9"
    summary_sheet.sheet_properties.tabColor = "D5CBE3"
    best_sheet.sheet_properties.tabColor = "BFD8BF"
    follow_up_sheet.sheet_properties.tabColor = "E8C9B8"
    no_shipments_sheet.sheet_properties.tabColor = "D8DDE6"


       # ======================================
    # Best Accounts
    # ======================================

    best_accounts_clean = []

    for row in best_accounts:

        if not isinstance(
            row,
            dict
        ):
            continue

        clean_row = row.copy()

        clean_row.pop(
            "Pending",
            None
        )

        best_accounts_clean.append(
            clean_row
        )

    # ======================================
    # Update Summary
    # ======================================

    summary_for_excel = dict(
        summary
    )

    summary_for_excel[
        "best_accounts"
    ] = len(
        best_accounts_clean
    )


    # ======================================
    # Export Summary
    # ======================================

    export_summary(
        summary_sheet,
        summary_for_excel
    )


    # ======================================
    # Export Best Accounts
    # ======================================

    export_sheet(
        best_sheet,

        add_ranking(
            best_accounts_clean
        ),

        "BEST ACCOUNTS",

        (
            "No accounts reached the selected "
            "delivery percentage and shipment criteria."
        )
        
    )

    # ======================================
    # Export Follow Up
    # ======================================

    export_sheet(
        follow_up_sheet,

        add_ranking(
            follow_up_accounts
        ),

        "NEED FOLLOW UP",

        "No accounts need follow up."
    )


    # ======================================
    # Export No Shipments
    # ======================================

    export_sheet(
        no_shipments_sheet,

        no_shipments,

        "NO SHIPMENTS",

        "All accounts created shipments."
    )


    # ======================================
    # Export All Businesses
    # ======================================

    export_sheet(
        all_accounts_sheet,

        all_accounts,

        f"ALL BUSINESSES - TOTAL: {len(all_accounts)}",

        "No businesses found."
    )


    # ======================================
    # Save
    # ======================================

    output_directory = Path(
        __file__
    ).resolve().parent

    try:
        start_value = datetime.strptime(str(start_date), "%Y-%m-%d")
        end_value = datetime.strptime(str(end_date), "%Y-%m-%d")
        if start_value.year == end_value.year and start_value.month == end_value.month:
            report_name = f"Sales Report For Month {start_value.strftime('%B')}.xlsx"
        else:
            report_name = f"Sales Report From {start_date} To {end_date}.xlsx"
    except Exception:
        report_name = f"Sales Report {datetime.now():%Y%m%d_%H%M%S}.xlsx"
    filename = output_directory / report_name

    workbook.save(
        filename
    )

    print()
    print("=" * 60)
    print("Excel Report Created Successfully")
    print(f"File: {filename}")
    print("=" * 60)

    return str(
        filename
    )